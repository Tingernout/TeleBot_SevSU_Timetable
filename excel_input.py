import openpyxl as excel
from openpyxl.utils import get_column_letter
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import openpyxl
def format_str_70(stroka):
    total= ''
    for i in range(len(stroka)):
        if i%70==0 and i != 0:
            total+='\n'
        total+=stroka[i]
    return total
def excel_to_png(day_name, df, png_file):
    import matplotlib.pyplot as plt
    import pandas as pd
    from matplotlib.table import Table


    # Настраиваем фигуру
    fig, ax = plt.subplots(figsize=(16, 9))
    ax.axis('tight')
    ax.axis('off')
    ax.set_title(day_name)

    # Создаем таблицу
    tbl = Table(ax, bbox=[0, 0, 1, 1])

    # Заполняем таблицу данными
    n_rows, n_cols = df.shape
    col_widths = [max(len(max(str(val).split('\n'))) for val in df[col]) for col in df.columns]
    row_heights = [max(len(str(val).split('\n')) for val in row) for row in df.itertuples(index=False, name=None)]

    # Добавляем ячейки в таблицу
    for i, col in enumerate(df.columns):
        tbl.add_cell(0, i, width=col_widths[i] / 10, height=0.3, text=col, loc='center', facecolor='grey')
    print(df)
    for i in range(n_rows):
        for j in range(n_cols):
            cell_value = str(df.iat[i, j])
            cell_height = row_heights[i]

            # Проверяем, является ли текущая ячейка объединенной
            if i > 0 and cell_value == 'nan':
                tbl.add_cell(i + 1, j, width=col_widths[j] / 10, height=0.2 * cell_height, text="", loc='center')
            else:
                cell_value = format_str_70(cell_value)
                tbl.add_cell(i + 1, j, width=col_widths[j] / 10, height=0.2 * cell_height, text=cell_value,
                             loc='center')

    # Настраиваем ширину и высоту таблицы
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(10)
    ax.add_table(tbl)

    # Сохраняем таблицу в файл
    plt.savefig(png_file, bbox_inches='tight')
    plt.show()
def width_column_well(worksheet,data):
    column_widths = [3,7,40,40,6,9,40,40,6,9]
    # for row in data:
    #     i=0
    #     for cell in row:
    #         if len(column_widths) > i:
    #             if len(str(cell)) > column_widths[i]:
    #                 column_widths[i] = len(str(cell))
    #         else:
    #             column_widths += [len(str(cell))]
    #         i+=1
    print(data)
    for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
        worksheet.column_dimensions[get_column_letter(i)].width = column_width





# Функция для разбиения расписания с главного сайта на маленькие части
# def data_in_excel(data):
#     fl = excel.load_workbook("temp.xlsx")
#     # fl = excel.Workbook()
#     week_id = 1
#     cur = 2
#     temp_data=[]
#     for day in data:
#         sheet = fl[f"PartWeek{week_id}"]
#         # sheet = fl.active
#         cur +=3
#         for lesson in day[1:]:
#             if lesson != []:
#                 cur_car = "C"
#                 for el in lesson[2:]:
#                     sheet[cur_car+str(cur)] = format_str_35(el.replace('\n',''))
#                     cur_car = str(chr(ord(cur_car) + 1))
#                 temp_data+=[lesson]
#             cur+=1
#         cur+=1
#         if cur>36:
#             cur=2
#             week_id=2
#             temp_data=[]
#         width_column_well(sheet,temp_data)
#     fl.save("rep.xlsx")
excel_file = 'Понедельник (24.12.2024)'  # Замените на имя вашего Excel файла

file_path = 'rep.xlsx'
df = pd.read_excel(file_path)
png_file = 'table.png'       # Замените на имя файла PNG

excel_to_png(excel_file, df, png_file)